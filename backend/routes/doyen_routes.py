import os
from io import BytesIO
from html import escape

from flask import Blueprint, jsonify, request, send_file, Response

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)

from database.reinitialisation_repository import (
    get_resume_reinitialisation_annuelle,
    creer_demande_reinitialisation,
    executer_reinitialisation_annuelle
)

from database.doyen_repository import (
    get_doyen_id_by_utilisateur_id,
    get_fiches_en_attente,
    get_fiche_complete,
    valider_ou_reviser_fiche,
    rechercher_fiches_par_id_universitaire
)

from database.chat_orientation_repository import (
    lister_documents_a_confirmer,
    confirmer_ou_refuser_document,
    get_document_info,
    get_statistiques_places_filieres,
    generer_contenu_txt_discussion,
    get_discussion_chat_par_fiche,
    get_choix_final_details,
    get_historique_etudiant_par_fiche,
    get_tableau_bord_doyen_avance,
    get_suivi_promotion_doyen,
    importer_liste_officielle_promotion,
    get_dossier_archives_exports_excel,
    normaliser_nom_archive,
    enregistrer_archive_export_excel,
    lister_archives_administratives,
    get_archive_fiche_engagement,
    get_archive_export_excel
)


try:
    from services.email_notification_service import email_configure
except Exception:
    def email_configure():
        return False
from services.jwt_service import doyen_requis, get_utilisateur_connecte_id
from database.notification_repository import (
    notifier_etudiant_decision_document
)


from services.annee_universitaire_service import (
    assurer_annee_universitaire_active,
    lister_annees_universitaires
)

doyen_bp = Blueprint(
    "doyen",
    __name__,
    url_prefix="/api/doyen"
)


def nettoyer_texte_pdf(texte):
    if texte is None:
        return ""

    return escape(str(texte)).replace("\n", "<br/>")


def formater_role_message_pdf(role_message):
    if role_message == "assistant":
        return "Chatbot"

    if role_message == "etudiant":
        return "Étudiant"

    if role_message == "user":
        return "Étudiant"

    if role_message == "bot":
        return "Chatbot"

    return role_message or "Message"


def recuperer_date_message_pdf(message):
    return (
        message.get("date_creation")
        or message.get("date_message")
        or message.get("created_at")
        or message.get("date_envoi")
        or ""
    )


def generer_pdf_discussion_buffer(discussion):
    fiche = discussion["fiche"]
    session_chat = discussion["session_chat"]
    scores = discussion["scores"]
    messages = discussion["messages"]

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.7 * cm,
        leftMargin=1.7 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm
    )

    styles = getSampleStyleSheet()

    titre_style = ParagraphStyle(
        "TitrePrincipal",
        parent=styles["Title"],
        fontSize=18,
        leading=22,
        alignment=1,
        textColor=colors.HexColor("#1d4ed8"),
        spaceAfter=16
    )

    sous_titre_style = ParagraphStyle(
        "SousTitre",
        parent=styles["Heading2"],
        fontSize=12,
        leading=15,
        textColor=colors.HexColor("#1f2937"),
        spaceBefore=10,
        spaceAfter=8
    )

    normal_style = ParagraphStyle(
        "TexteNormal",
        parent=styles["Normal"],
        fontSize=9,
        leading=12,
        spaceAfter=5
    )

    message_style = ParagraphStyle(
        "MessageStyle",
        parent=styles["Normal"],
        fontSize=9,
        leading=12,
        leftIndent=8,
        rightIndent=8,
        spaceBefore=4,
        spaceAfter=6
    )

    elements = []

    elements.append(Paragraph("Trace de discussion - Chatbot d’orientation", titre_style))

    elements.append(Paragraph("Informations étudiant", sous_titre_style))

    infos_etudiant = [
        ["Nom complet", f"{fiche.get('prenom', '')} {fiche.get('nom', '')}"],
        ["ID universitaire", fiche.get("id_universitaire", "")],
        ["Email", fiche.get("email", "")],
        ["Promotion", fiche.get("promotion", "")]
    ]

    table_infos = Table(infos_etudiant, colWidths=[5 * cm, 11 * cm])
    table_infos.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eff6ff")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#1e3a8a")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6)
    ]))

    elements.append(table_infos)
    elements.append(Spacer(1, 8))

    elements.append(Paragraph("Informations orientation", sous_titre_style))

    infos_orientation = [
        ["Fiche ID", fiche.get("fiche_id", "")],
        ["Test orientation ID", fiche.get("test_orientation_id", "")],
        ["Spécialité proposée", fiche.get("specialite_recommandee", "")],
        ["Date génération", fiche.get("date_generation", "")]
    ]

    table_orientation = Table(infos_orientation, colWidths=[5 * cm, 11 * cm])
    table_orientation.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f8fafc")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#111827")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6)
    ]))

    elements.append(table_orientation)
    elements.append(Spacer(1, 8))

    elements.append(Paragraph("Scores par filière", sous_titre_style))

    if len(scores) == 0:
        elements.append(Paragraph("Aucun score enregistré.", normal_style))
    else:
        donnees_scores = [["Filière", "Score", "Pourcentage"]]

        for score in scores:
            donnees_scores.append([
                score.get("specialite", ""),
                str(score.get("score", "")),
                f"{score.get('pourcentage', '')}%"
            ])

        table_scores = Table(donnees_scores, colWidths=[9 * cm, 3 * cm, 4 * cm])
        table_scores.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1d4ed8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6)
        ]))

        elements.append(table_scores)

    elements.append(Spacer(1, 8))

    elements.append(Paragraph("Session chatbot", sous_titre_style))

    if session_chat is None:
        elements.append(Paragraph("Aucune session chat trouvée pour cette fiche.", normal_style))
    else:
        elements.append(Paragraph(
            f"<b>Session ID :</b> {session_chat.get('id', '')}<br/>"
            f"<b>Statut :</b> {session_chat.get('statut', '')}<br/>"
            f"<b>Raison droit :</b> {session_chat.get('raison_droit', '')}",
            normal_style
        ))

    elements.append(Spacer(1, 8))

    elements.append(Paragraph("Discussion complète", sous_titre_style))

    if len(messages) == 0:
        elements.append(Paragraph("Aucun message enregistré pour cette discussion.", normal_style))
    else:
        for index, message in enumerate(messages, start=1):
            role = formater_role_message_pdf(message.get("role_message"))
            date_message = recuperer_date_message_pdf(message)
            contenu = nettoyer_texte_pdf(message.get("contenu") or "")

            elements.append(Paragraph(
                f"<b>Message {index} - {role}</b>",
                normal_style
            ))

            if date_message:
                elements.append(Paragraph(
                    f"<b>Date :</b> {nettoyer_texte_pdf(date_message)}",
                    normal_style
                ))

            elements.append(Paragraph(contenu, message_style))
            elements.append(Spacer(1, 4))

    document.build(elements)

    buffer.seek(0)
    return buffer





def nettoyer_nom_feuille_excel(nom):
    nom = str(nom or "Feuille").strip()

    caracteres_interdits = ["\\", "/", "?", "*", "[", "]", ":"]

    for caractere in caracteres_interdits:
        nom = nom.replace(caractere, "-")

    if len(nom) == 0:
        nom = "Feuille"

    return nom[:31]


def ecrire_ligne_excel(ws, ligne_index, valeurs, style_entete=None):
    for colonne_index, valeur in enumerate(valeurs, start=1):
        cellule = ws.cell(row=ligne_index, column=colonne_index, value=valeur)

        if style_entete:
            cellule.font = style_entete["font"]
            cellule.fill = style_entete["fill"]
            cellule.alignment = style_entete["alignment"]
            cellule.border = style_entete["border"]



def ajuster_largeurs_excel(ws):
    for colonne in ws.columns:
        lettre = colonne[0].column_letter
        largeur = 12

        for cellule in colonne:
            valeur = cellule.value

            if valeur is not None:
                largeur = max(largeur, min(len(str(valeur)) + 2, 38))

        ws.column_dimensions[lettre].width = largeur


def generer_excel_suivi_promotion(resultat_suivi):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    suivi = resultat_suivi.get("suivi", [])

    titre_font = Font(bold=True, size=16, color="1E3A8A")
    sous_titre_font = Font(bold=True, size=11, color="475569")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_style = {
        "font": header_font,
        "fill": header_fill,
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": border
    }

    headers = [
        "Nom",
        "Prénom",
        "ID universitaire",
        "Email Outlook"
    ]

    filieres = sorted({
        ligne.get("filiere_choisie")
        for ligne in suivi
        if ligne.get("filiere_choisie") and ligne.get("filiere_choisie") != "Non choisie"
    })

    if len(filieres) == 0:
        filieres = ["Aucune filière"]

    premiere_feuille = True

    def ajouter_feuille_filiere(nom_filiere, lignes):
        nonlocal premiere_feuille

        if premiere_feuille:
            ws = wb.active
            ws.title = nettoyer_nom_feuille_excel(nom_filiere)
            premiere_feuille = False
        else:
            ws = wb.create_sheet(nettoyer_nom_feuille_excel(nom_filiere))

        ws.merge_cells("A1:D1")
        ws["A1"] = nom_filiere
        ws["A1"].font = titre_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("A2:D2")
        ws["A2"] = "Liste des étudiants inscrits dans cette filière"
        ws["A2"].font = sous_titre_font
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

        ecrire_ligne_excel(ws, 4, headers, header_style)
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:D{max(4, len(lignes) + 4)}"

        for row_index, etudiant in enumerate(lignes, start=5):
            valeurs = [
                etudiant.get("nom", ""),
                etudiant.get("prenom", ""),
                etudiant.get("id_universitaire", ""),
                etudiant.get("email_outlook", "") or etudiant.get("email_plateforme", "")
            ]

            for col_index, valeur in enumerate(valeurs, start=1):
                cellule = ws.cell(row=row_index, column=col_index, value=valeur)
                cellule.border = border
                cellule.alignment = Alignment(vertical="top", wrap_text=True)

        ajuster_largeurs_excel(ws)
        ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 18)
        ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width, 18)
        ws.column_dimensions["C"].width = max(ws.column_dimensions["C"].width, 20)
        ws.column_dimensions["D"].width = max(ws.column_dimensions["D"].width, 34)

    for filiere in filieres:
        lignes_filiere = [
            ligne for ligne in suivi
            if ligne.get("filiere_choisie") == filiere
        ]

        ajouter_feuille_filiere(filiere, lignes_filiere)

    return wb


@doyen_bp.route("/tableau-bord/avance", methods=["GET"])
@doyen_requis
def tableau_bord_avance():
    resultat = get_tableau_bord_doyen_avance()

    status_code = 200 if resultat.get("success") else 500

    return jsonify(resultat), status_code



@doyen_bp.route("/promotion/suivi", methods=["GET"])
@doyen_requis
def suivi_promotion_doyen():
    resultat = get_suivi_promotion_doyen()

    status_code = 200 if resultat.get("success") else 400

    return jsonify(resultat), status_code


@doyen_bp.route("/promotion/liste-officielle/importer", methods=["POST"])
@doyen_requis
def importer_liste_officielle_doyen():
    data = request.get_json()

    if data is None:
        return jsonify({
            "success": False,
            "message": "Aucune donnée reçue."
        }), 400

    texte_csv = data.get("texte_csv", "")
    etudiants = data.get("etudiants")
    remplacer = bool(data.get("remplacer", False))

    resultat = importer_liste_officielle_promotion(
        texte_csv=texte_csv,
        etudiants=etudiants,
        remplacer=remplacer
    )

    status_code = 200 if resultat.get("success") else 400

    return jsonify(resultat), status_code


@doyen_bp.route("/promotion/export-excel", methods=["GET"])
@doyen_requis
def exporter_suivi_promotion_excel():
    try:
        resultat = get_suivi_promotion_doyen()

        if not resultat.get("success"):
            return jsonify(resultat), 400

        try:
            wb = generer_excel_suivi_promotion(resultat)
        except ModuleNotFoundError:
            return jsonify({
                "success": False,
                "message": (
                    "Le module openpyxl n’est pas installé. "
                    "Installez-le avec : pip install openpyxl"
                )
            }), 500

        annee_universitaire = resultat.get("annee_universitaire", "promotion")
        annee = str(annee_universitaire).replace("/", "_")
        nom_fichier = f"suivi_promotion_{annee}.xlsx"

        dossier_archive = os.path.join(
            get_dossier_archives_exports_excel(),
            normaliser_nom_archive(annee_universitaire)
        )
        os.makedirs(dossier_archive, exist_ok=True)

        chemin_archive = os.path.join(dossier_archive, nom_fichier)
        wb.save(chemin_archive)
        enregistrer_archive_export_excel(
            annee_universitaire,
            nom_fichier,
            chemin_archive
        )

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=nom_fichier,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as erreur:
        return jsonify({
            "success": False,
            "message": "Erreur lors de la génération de l’export Excel.",
            "erreur": str(erreur)
        }), 500


@doyen_bp.route("/archives/administratives", methods=["GET"])
@doyen_requis
def archives_administratives_doyen():
    resultat = lister_archives_administratives()
    status_code = 200 if resultat.get("success") else 500
    return jsonify(resultat), status_code


@doyen_bp.route("/archives/fiches/<int:archive_id>/visualiser", methods=["GET"])
@doyen_requis
def visualiser_archive_fiche_engagement(archive_id):
    archive = get_archive_fiche_engagement(archive_id)

    if archive is None:
        return jsonify({
            "success": False,
            "message": "Archive introuvable."
        }), 404

    chemin = archive.get("chemin_archive")

    if not chemin or not os.path.exists(chemin):
        return jsonify({
            "success": False,
            "message": "Fichier d’archive introuvable sur le serveur."
        }), 404

    return send_file(
        chemin,
        as_attachment=False,
        download_name=archive.get("nom_fichier_archive") or os.path.basename(chemin)
    )


@doyen_bp.route("/archives/fiches/<int:archive_id>/telecharger", methods=["GET"])
@doyen_requis
def telecharger_archive_fiche_engagement(archive_id):
    archive = get_archive_fiche_engagement(archive_id)

    if archive is None:
        return jsonify({
            "success": False,
            "message": "Archive introuvable."
        }), 404

    chemin = archive.get("chemin_archive")

    if not chemin or not os.path.exists(chemin):
        return jsonify({
            "success": False,
            "message": "Fichier d’archive introuvable sur le serveur."
        }), 404

    return send_file(
        chemin,
        as_attachment=True,
        download_name=archive.get("nom_fichier_archive") or os.path.basename(chemin)
    )


@doyen_bp.route("/archives/exports/<int:archive_id>/telecharger", methods=["GET"])
@doyen_requis
def telecharger_archive_export_excel(archive_id):
    archive = get_archive_export_excel(archive_id)

    if archive is None:
        return jsonify({
            "success": False,
            "message": "Export introuvable."
        }), 404

    chemin = archive.get("chemin_archive")

    if not chemin or not os.path.exists(chemin):
        return jsonify({
            "success": False,
            "message": "Fichier Excel introuvable sur le serveur."
        }), 404

    return send_file(
        chemin,
        as_attachment=True,
        download_name=archive.get("nom_fichier") or os.path.basename(chemin),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@doyen_bp.route("/annee-universitaire/active", methods=["GET"])
@doyen_requis
def annee_universitaire_active():
    resultat = assurer_annee_universitaire_active()

    status_code = 200 if resultat.get("success") else 500

    return jsonify(resultat), status_code


@doyen_bp.route("/annees-universitaires", methods=["GET"])
@doyen_requis
def annees_universitaires():
    resultat = lister_annees_universitaires()

    status_code = 200 if resultat.get("success") else 500

    return jsonify(resultat), status_code


@doyen_bp.route("/test", methods=["GET"])
def test_doyen():
    return jsonify({
        "message": "Routes doyen fonctionnelles."
    })


@doyen_bp.route("/fiches/en-attente", methods=["GET"])
@doyen_requis
def fiches_en_attente():
    fiches = get_fiches_en_attente()

    return jsonify({
        "success": True,
        "fiches": fiches
    })


@doyen_bp.route("/fiches/rechercher/<recherche>", methods=["GET"])
@doyen_requis
def rechercher_fiches(recherche):
    fiches = rechercher_fiches_par_id_universitaire(recherche)

    return jsonify({
        "success": True,
        "fiches": fiches
    })


@doyen_bp.route("/fiches/<int:fiche_id>", methods=["GET"])
@doyen_requis
def fiche_complete(fiche_id):
    fiche = get_fiche_complete(fiche_id)

    if fiche is None:
        return jsonify({
            "success": False,
            "message": "Fiche introuvable."
        }), 404

    return jsonify({
        "success": True,
        "fiche": fiche
    })


@doyen_bp.route("/fiches/<int:fiche_id>/historique", methods=["GET"])
@doyen_requis
def historique_etudiant_fiche(fiche_id):
    resultat = get_historique_etudiant_par_fiche(fiche_id)

    if not resultat.get("success"):
        return jsonify(resultat), 404

    return jsonify(resultat), 200


@doyen_bp.route("/fiches/<int:fiche_id>/discussion", methods=["GET"])
@doyen_requis
def visualiser_discussion_fiche(fiche_id):
    resultat = get_discussion_chat_par_fiche(fiche_id)

    if not resultat.get("success"):
        return jsonify(resultat), 404

    return jsonify(resultat), 200


@doyen_bp.route("/fiches/<int:fiche_id>/discussion/txt", methods=["GET"])
@doyen_requis
def telecharger_discussion_txt(fiche_id):
    resultat = generer_contenu_txt_discussion(fiche_id)

    if not resultat.get("success"):
        return jsonify(resultat), 404

    contenu = resultat["contenu"]
    nom_fichier = resultat["nom_fichier"]

    return Response(
        contenu,
        mimetype="text/plain; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename={nom_fichier}"
        }
    )


@doyen_bp.route("/fiches/<int:fiche_id>/discussion/pdf", methods=["GET"])
@doyen_requis
def telecharger_discussion_pdf(fiche_id):
    resultat = get_discussion_chat_par_fiche(fiche_id)

    if not resultat.get("success"):
        return jsonify(resultat), 404

    fiche = resultat["fiche"]

    nom_fichier = (
        f"discussion_chat_"
        f"{fiche.get('id_universitaire', 'etudiant')}_"
        f"fiche_{fiche_id}.pdf"
    )

    buffer_pdf = generer_pdf_discussion_buffer(resultat)

    return send_file(
        buffer_pdf,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=nom_fichier
    )


@doyen_bp.route("/fiches/<int:fiche_id>/valider", methods=["POST"])
@doyen_requis
def valider_fiche(fiche_id):
    data = request.get_json()

    if data is None:
        return jsonify({
            "success": False,
            "message": "Aucune donnée reçue."
        }), 400

    utilisateur_id = data.get("utilisateur_id")
    statut = data.get("statut")
    remarque = data.get("remarque")

    resultat = valider_ou_reviser_fiche(
        fiche_id,
        utilisateur_id,
        statut,
        remarque
    )

    status_code = 200 if resultat["success"] else 400

    return jsonify(resultat), status_code


@doyen_bp.route("/autoriser-nouveau-test", methods=["POST"])
@doyen_requis
def autoriser_nouveau_test():
    return jsonify({
        "success": False,
        "message": (
            "Cette action n’est plus disponible. "
            "L’étudiant peut repasser le test librement depuis son espace."
        )
    }), 403


@doyen_bp.route("/repartition-choix-filieres", methods=["GET"])
@doyen_requis
def repartition_choix_filieres():
    repartition = get_statistiques_places_filieres()

    return jsonify({
        "success": True,
        "repartition": repartition
    })


@doyen_bp.route("/places", methods=["GET"])
@doyen_requis
def statistiques_places():
    # Ancienne route conservée uniquement pour compatibilité technique.
    # La logique métier affichée est maintenant la répartition des choix par filière.
    repartition = get_statistiques_places_filieres()

    return jsonify({
        "success": True,
        "statistiques": repartition
    })


@doyen_bp.route("/documents/a-confirmer", methods=["GET"])
@doyen_requis
def documents_a_confirmer():
    documents = lister_documents_a_confirmer()

    return jsonify({
        "success": True,
        "documents": documents,
        "total": len(documents)
    })


@doyen_bp.route("/documents/<int:document_id>/visualiser", methods=["GET"])
@doyen_requis
def visualiser_document(document_id):
    document = get_document_info(document_id)

    if document is None:
        return jsonify({
            "success": False,
            "message": "Document introuvable."
        }), 404

    chemin = document["chemin_fichier"]

    return send_file(
        chemin,
        as_attachment=False,
        download_name=document["nom_fichier_original"]
    )


@doyen_bp.route("/documents/decision", methods=["POST"])
@doyen_requis
def decision_document():
    data = request.get_json()

    if data is None:
        return jsonify({
            "success": False,
            "message": "Aucune donnée reçue."
        }), 400

    utilisateur_id = data.get("utilisateur_id")
    choix_id = data.get("choix_id")
    decision = data.get("decision")
    remarque = data.get("remarque")

    if utilisateur_id is None:
        return jsonify({
            "success": False,
            "message": "utilisateur_id obligatoire."
        }), 400

    if choix_id is None:
        return jsonify({
            "success": False,
            "message": "choix_id obligatoire."
        }), 400

    doyen_id = get_doyen_id_by_utilisateur_id(utilisateur_id)

    if doyen_id is None:
        return jsonify({
            "success": False,
            "message": "Aucun doyen trouvé pour cet utilisateur."
        }), 404

    resultat = confirmer_ou_refuser_document(
        choix_id,
        doyen_id,
        decision,
        remarque
    )

    if resultat.get("success"):
        details_choix = get_choix_final_details(choix_id)

        if details_choix is not None:
            notifier_etudiant_decision_document(
                details_choix,
                decision,
                remarque
            )

    status_code = 200 if resultat["success"] else 400

    return jsonify(resultat), status_code


@doyen_bp.route("/notifications/email/etat", methods=["GET"])
@doyen_requis
def etat_notification_email():
    configuration_active = email_configure()

    if configuration_active:
        message = "La configuration email SMTP est active."
    else:
        message = (
            "La configuration email SMTP n’est pas encore active. "
            "Les documents peuvent être déposés, mais les notifications email "
            "ne seront pas envoyées."
        )

    return jsonify({
        "success": True,
        "email_configure": configuration_active,
        "message": message
    }), 200


@doyen_bp.route("/reinitialisation/resume/<annee_universitaire>", methods=["GET"])
@doyen_requis
def resume_reinitialisation(annee_universitaire):
    resultat = get_resume_reinitialisation_annuelle(annee_universitaire)

    status_code = 200 if resultat["success"] else 400

    return jsonify(resultat), status_code


@doyen_bp.route("/reinitialisation/demande", methods=["POST"])
@doyen_requis
def demande_reinitialisation():
    data = request.get_json()

    if data is None:
        return jsonify({
            "success": False,
            "message": "Aucune donnée reçue."
        }), 400

    utilisateur_id = data.get("utilisateur_id")
    annee_universitaire = data.get("annee_universitaire")
    mot_de_passe = data.get("mot_de_passe")
    phrase_securite = data.get("phrase_securite")

    if utilisateur_id is None:
        return jsonify({
            "success": False,
            "message": "utilisateur_id obligatoire."
        }), 400

    if not annee_universitaire:
        return jsonify({
            "success": False,
            "message": "Année universitaire obligatoire."
        }), 400

    if not mot_de_passe:
        return jsonify({
            "success": False,
            "message": "Mot de passe obligatoire."
        }), 400

    if not phrase_securite:
        return jsonify({
            "success": False,
            "message": "Phrase de sécurité obligatoire."
        }), 400

    resultat = creer_demande_reinitialisation(
        utilisateur_id,
        annee_universitaire,
        mot_de_passe,
        phrase_securite
    )

    status_code = 200 if resultat["success"] else 400

    return jsonify(resultat), status_code


@doyen_bp.route("/reinitialisation/confirmer", methods=["POST"])
@doyen_requis
def confirmer_reinitialisation():
    data = request.get_json()

    if data is None:
        return jsonify({
            "success": False,
            "message": "Aucune donnée reçue."
        }), 400

    utilisateur_id = data.get("utilisateur_id")
    reinitialisation_id = data.get("reinitialisation_id")
    code_confirmation = data.get("code_confirmation")

    if utilisateur_id is None:
        return jsonify({
            "success": False,
            "message": "utilisateur_id obligatoire."
        }), 400

    if reinitialisation_id is None:
        return jsonify({
            "success": False,
            "message": "reinitialisation_id obligatoire."
        }), 400

    if not code_confirmation:
        return jsonify({
            "success": False,
            "message": "Code de confirmation obligatoire."
        }), 400

    resultat = executer_reinitialisation_annuelle(
        utilisateur_id,
        reinitialisation_id,
        code_confirmation
    )

    status_code = 200 if resultat["success"] else 400

    return jsonify(resultat), status_code
